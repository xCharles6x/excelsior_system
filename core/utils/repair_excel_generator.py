"""
core/utils/repair_excel_generator.py
─────────────────────────────────────
Generates an Excel workbook (.xlsx) with two sheets:
  Sheet 1 — Warranty Certificate
  Sheet 2 — Maintenance Service Report Form
Auto-filled from a RepairInspection model instance.

Usage (in views.py):
    from .utils.repair_excel_generator import generate_repair_excel
    buf = generate_repair_excel(record, logo_path, sig_path)
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImg


DARK_BLUE  = "1F3864"
LIGHT_BLUE = "D9E1F2"
WHITE      = "FFFFFF"
BLACK      = "000000"
GRAY       = "DDDDDD"


# ── Style helpers ─────────────────────────────────────────────────────────────

def _T(style='thin', color='000000'):
    return Side(style=style, color=color)

def _box(thick=False):
    s = _T('medium' if thick else 'thin')
    return Border(top=s, bottom=s, left=s, right=s)

def _bot():
    return Border(bottom=_T('thin'))

def _hline(ws, row, c1, c2, style='medium'):
    s = _T(style)
    for c in range(c1, c2 + 1):
        cell = ws.cell(row=row, column=c)
        old  = cell.border
        cell.border = Border(top=old.top, bottom=s, left=old.left, right=old.right)

def _sc(ws, row, col, val='', bold=False, sz=10, bg=None, color=BLACK,
        ha='left', va='center', bdr=None, wrap=False, italic=False, ul=False):
    cell = ws.cell(row=row, column=col)
    cell.value = val
    cell.font  = Font(name='Arial', bold=bold, size=sz, color=color,
                      italic=italic, underline='single' if ul else None)
    cell.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
    if bg:  cell.fill   = PatternFill('solid', start_color=bg)
    if bdr: cell.border = bdr

def _mg(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

def _cw(ws, col, w):
    ws.column_dimensions[get_column_letter(col)].width = w

def _rh(ws, row, h):
    ws.row_dimensions[row].height = h

def _outer(ws, r1, r2, c1, c2):
    for r in range(r1, r2 + 1):
        for col in [c1, c2]:
            cell = ws.cell(row=r, column=col)
            old  = cell.border
            if col == c1:
                cell.border = Border(top=old.top, bottom=old.bottom,
                                     left=_T('medium'), right=old.right)
            else:
                cell.border = Border(top=old.top, bottom=old.bottom,
                                     left=old.left, right=_T('medium'))
    for c in range(c1, c2 + 1):
        for row in [r1, r2]:
            cell = ws.cell(row=row, column=c)
            old  = cell.border
            if row == r1:
                cell.border = Border(top=_T('medium'), bottom=old.bottom,
                                     left=old.left, right=old.right)
            else:
                cell.border = Border(top=old.top, bottom=_T('medium'),
                                     left=old.left, right=old.right)


# ── Main generator ────────────────────────────────────────────────────────────

def generate_repair_excel(record, logo_path=None, sig_path=None):
    """
    Args:
        record    : RepairInspection model instance
        logo_path : absolute path to logo image file (optional)
        sig_path  : absolute path to signature image file (optional)
    Returns:
        io.BytesIO buffer ready for HttpResponse
    """
    wb = Workbook()
    _build_warranty(wb.active, record, logo_path, sig_path)
    ws2 = wb.create_sheet("Maintenance Service Report")
    _build_msr(ws2, record, logo_path, sig_path)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Sheet 1: Warranty Certificate ────────────────────────────────────────────

def _build_warranty(ws, r, logo_path, sig_path):
    ws.title = "Warranty Certificate"

    for i, w in enumerate([1,11,10,10,10,10,10,10,10,1], 1):
        _cw(ws, i, w)
    for row in range(1, 36):
        _rh(ws, row, 15)

    # ── Header
    _rh(ws,2,30); _rh(ws,3,13); _rh(ws,4,13); _rh(ws,5,13)
    if logo_path:
        try:
            logo = XLImg(logo_path)
            logo.width = 85; logo.height = 85; logo.anchor = 'B2'
            ws.add_image(logo)
        except Exception:
            pass

    _mg(ws,2,3,2,9); _sc(ws,2,3,'Excelsior Aviation GSE Corporation',True,14,va='bottom')
    _mg(ws,3,3,3,9); _sc(ws,3,3,'Lot 9 Dona Lucing, Calibutbut Bacolor 2001 Pampanga Philippines',sz=9)
    _mg(ws,4,3,4,9); _sc(ws,4,3,'Mobile No. +63 905 463 7252  +63 917 158 8844',sz=9)
    _mg(ws,5,3,5,9); _sc(ws,5,3,'sale.service@excelsioraviation.com     www.excelsioraviation.com',sz=9)
    _rh(ws,6,4); _hline(ws,6,2,9,'medium')

    # ── Cert No
    _rh(ws,8,14)
    year = str(r.date.year) if r.date else '2026'
    _mg(ws,8,2,8,9); _sc(ws,8,2,f'No. WC{year}-XXXX',True,11,ha='right',ul=True)

    # ── Title
    _rh(ws,9,26)
    _mg(ws,9,2,9,9); _sc(ws,9,2,'WARRANTY CERTIFICATE',True,16,ha='center',va='center')
    _rh(ws,10,4); _hline(ws,10,2,9,'thin')
    _rh(ws,11,6)

    # ── Fields
    for row in range(12,19): _rh(ws,row,16)

    _sc(ws,12,2,'CUSTOMER',True,10);  _sc(ws,12,3,':',True,10)
    _mg(ws,12,4,12,5); _sc(ws,12,4,r.customer_name or '',bdr=_bot())
    _sc(ws,12,6,'ADDRESS',True,10);   _sc(ws,12,7,':',True,10)
    _mg(ws,12,8,12,9); _sc(ws,12,8,r.customer_address or '',bdr=_bot(),wrap=True)

    _sc(ws,13,2,'DATE OF REPAIR',True,10); _sc(ws,13,3,':',True,10)
    _mg(ws,13,4,13,5); _sc(ws,13,4,str(r.date) if r.date else '',bdr=_bot())
    _sc(ws,13,6,'WARRANTY PERIOD',True,10); _sc(ws,13,7,':',True,10)
    _mg(ws,13,8,13,9); _sc(ws,13,8,'2 Years',bdr=_bot())

    _sc(ws,14,2,'ITEM / DESCRIPTION',True,10); _sc(ws,14,3,':',True,10)
    _mg(ws,14,4,14,9); _sc(ws,14,4,r.description or '',bdr=_bot())

    _sc(ws,15,2,'MODEL NUMBER',True,10); _sc(ws,15,3,':',True,10)
    _mg(ws,15,4,15,5); _sc(ws,15,4,r.model_number or '',bdr=_bot())
    _sc(ws,15,6,'REPORT NO.',True,10); _sc(ws,15,7,':',True,10)
    _mg(ws,15,8,15,9); _sc(ws,15,8,r.report_number or '',bdr=_bot())

    _sc(ws,16,2,'SERIAL NUMBER',True,10); _sc(ws,16,3,':',True,10)
    _mg(ws,16,4,16,5); _sc(ws,16,4,r.serial_number or '',bdr=_bot())
    _sc(ws,16,6,'MECHANIC',True,10); _sc(ws,16,7,':',True,10)
    _mg(ws,16,8,16,9); _sc(ws,16,8,r.mechanic or '',bdr=_bot())

    # ── Warranty statement
    _rh(ws,18,14); _mg(ws,18,2,18,9)
    _sc(ws,18,2,'WARRANTY STATEMENT',True,11)

    desc = r.description or 'The described equipment'
    _rh(ws,19,55); _mg(ws,19,2,19,9)
    _sc(ws,19,2,
        f'{desc} has been repaired and is fully warranted against Excelsior technician faults '
        f'and defective parts under our service warranty guide for a period of 2 Years from the '
        f'date of repair and delivery or upon completion of installation.',
        sz=10, wrap=True, va='top')

    # ── Exceptions
    _rh(ws,21,14); _mg(ws,21,2,21,9)
    _sc(ws,21,2,'WARRANTY EXCEPTIONS:',True,10,LIGHT_BLUE,BLACK)

    exceptions = [
        '1.  Damage resulting from accidents, misuse, abuse, over loading and failure to follow normal operating procedures.',
        '2.  Normal wear-and-tear, corrosion, rusting or stains.',
        '3.  Defects from improper testing, operation, maintenance, installation, or any alteration or modification of any kind.',
        '4.  General maintenance made other than Excelsior Aviation GSE Corporation authorized technician.',
        '5.  If any parts are replaced (third party brand) not supplied or approved by Excelsior, or unit dismantled by other technician.',
    ]
    for i, exc in enumerate(exceptions):
        row = 22 + i
        _rh(ws, row, 22 if len(exc) > 80 else 16)
        _mg(ws, row, 2, row, 9)
        _sc(ws, row, 2, exc, sz=10, bdr=_box(), wrap=True, va='top')

    # ── Signature
    _rh(ws,28,45); _rh(ws,29,14); _rh(ws,30,14)
    if sig_path:
        try:
            sig = XLImg(sig_path)
            sig.width = 120; sig.height = 45; sig.anchor = 'B28'
            ws.add_image(sig)
        except Exception:
            pass
    _mg(ws,28,6,28,9); _sc(ws,28,6,str(r.date) if r.date else '',sz=10,ha='right',va='bottom')
    _mg(ws,29,2,29,4); _sc(ws,29,2,r.mechanic or '',sz=10,ul=True)
    _mg(ws,29,6,29,9); _sc(ws,29,6,'Date',sz=10,ha='right')
    _mg(ws,30,2,30,4); _sc(ws,30,2,'Technician',sz=10,italic=True)

    # ── Notes
    _rh(ws,32,40); _mg(ws,32,2,32,9)
    _sc(ws,32,2,
        'Excelsior Aviation GSE Corporation shall not be liable for any incidents due to human error. '
        'Certificate is not valid without Excelsior dry seal.',
        sz=9, italic=True, wrap=True, va='top')

    # ── Footer
    _hline(ws,33,2,9,'thin')
    _rh(ws,34,14); _mg(ws,34,2,34,9)
    _sc(ws,34,2,'www.excelsioraviation.com',sz=9,ha='center',italic=True)
    _outer(ws,1,34,2,9)


# ── Sheet 2: Maintenance Service Report ──────────────────────────────────────

def _build_msr(ws, r, logo_path, sig_path):
    for i, w in enumerate([1,22,32,10,10,14,1], 1):
        _cw(ws, i, w)
    for row in range(1, 66):
        _rh(ws, row, 16)

    # ── Header
    _rh(ws,2,30); _rh(ws,3,13); _rh(ws,4,13); _rh(ws,5,13)
    if logo_path:
        try:
            logo = XLImg(logo_path)
            logo.width = 85; logo.height = 85; logo.anchor = 'B2'
            ws.add_image(logo)
        except Exception:
            pass

    _mg(ws,2,3,2,6); _sc(ws,2,3,'Excelsior Aviation GSE Corporation',True,14,va='bottom')
    _mg(ws,3,3,3,6); _sc(ws,3,3,'Lot 9 Dona Lucing, Calibutbut Bacolor 2001 Pampanga Philippines',sz=9)
    _mg(ws,4,3,4,6); _sc(ws,4,3,'Mobile No. +63 905 463 7252',sz=9)
    _mg(ws,5,3,5,6); _sc(ws,5,3,'sale.service@excelsioraviation.com     www.excelsioraviation.com',sz=9)
    _rh(ws,6,4); _hline(ws,6,2,6,'medium')
    _rh(ws,7,6)

    # ── Title
    _rh(ws,8,24); _mg(ws,8,2,8,6)
    _sc(ws,8,2,'MAINTENANCE SERVICE REPORT FORM',True,14,ha='center',va='center')
    _rh(ws,9,22); _mg(ws,9,2,9,6)
    _sc(ws,9,2,f'MSR{r.report_number or ""}',True,12,ha='center',va='center',bdr=_box(True))
    _rh(ws,10,8)

    # ── Customer
    for row, lbl, val in [(11,'Customer Name',r.customer_name),(12,'Address',r.customer_address)]:
        _rh(ws,row,18)
        _sc(ws,row,2,lbl,True,10,GRAY,bdr=_box())
        _mg(ws,row,3,row,6); _sc(ws,row,3,val or '',sz=10,bdr=_box(),wrap=True)
    _rh(ws,13,8)

    # ── Equipment
    _rh(ws,14,20); _mg(ws,14,2,14,6)
    _sc(ws,14,2,'Equipment Information',True,10,LIGHT_BLUE,BLACK,'center',bdr=_box())
    for row, lbl, val in [
        (15,'Equipment Description', r.description),
        (16,'Part No.',              r.model_number),
        (17,'Serial No.',            r.serial_number),
        (18,'Record Type',           r.get_record_type_display()),
    ]:
        _rh(ws,row,18)
        _sc(ws,row,2,lbl,True,10,GRAY,bdr=_box())
        _mg(ws,row,3,row,6); _sc(ws,row,3,val or '',sz=10,bdr=_box())
    _rh(ws,19,8)

    # ── Description of Service
    _rh(ws,20,20); _mg(ws,20,2,20,6)
    _sc(ws,20,2,'Description of Service',True,11,ha='center',bdr=_box())
    svc = ['Inspection','Repair','Load Test']
    rt  = (r.record_type or '').lower()
    for i, item in enumerate(svc):
        check = '☑' if item.lower() == rt else '☐'
        _rh(ws,21+i,16); _mg(ws,21+i,2,21+i,6)
        _sc(ws,21+i,2,f'{check}  {item}',sz=10,bdr=_box())
    _rh(ws,24,8)

    # ── Customer Report
    _rh(ws,25,20); _mg(ws,25,2,25,6)
    _sc(ws,25,2,'Customer Report / Problem Found',True,11,ha='center',bdr=_box())
    _rh(ws,26,55); _mg(ws,26,2,26,6)
    _sc(ws,26,2,r.customer_report or '',sz=10,bdr=_box(),wrap=True,va='top')
    _rh(ws,27,8)

    # ── Corrective Action
    _rh(ws,28,20); _mg(ws,28,2,28,6)
    _sc(ws,28,2,'Describe Corrective Action Completed or Required',True,11,ha='center',bdr=_box())
    lines = (r.diagnose_result or '').split('\n') if r.diagnose_result else []
    for i in range(10):
        row = 29 + i; _rh(ws,row,18); _mg(ws,row,2,row,6)
        line = lines[i] if i < len(lines) else ''
        txt  = f'• {line}' if line.strip() and not line.startswith('•') else line
        _sc(ws,row,2,txt,sz=10,bdr=_box(),wrap=True)
    _rh(ws,39,8)

    # ── Remarks
    _rh(ws,40,20); _mg(ws,40,2,40,6)
    _sc(ws,40,2,'Remarks',True,11,ha='center',bdr=_box())
    _rh(ws,41,22); _mg(ws,41,2,41,6)
    _sc(ws,41,2,r.remarks or '',sz=10,bdr=_box(),wrap=True)
    for row in [42,43]:
        _rh(ws,row,18); _mg(ws,row,2,row,6); _sc(ws,row,2,'',bdr=_box())
    _rh(ws,44,8)

    # ── Parts Installed
    _rh(ws,45,20); _mg(ws,45,2,45,6)
    _sc(ws,45,2,'Parts Installed',True,11,ha='center',bdr=_box())
    for col, hdr in zip([2,3,4,5,6],['Qty','Unit','Description','Part Number','']):
        _sc(ws,46,col,hdr,True,10,GRAY,ha='center',bdr=_box())
    for i in range(6):
        for col in range(2,7): _sc(ws,47+i,col,'',bdr=_box())
    _rh(ws,53,8)

    # ── Photos
    _rh(ws,54,20); _mg(ws,54,2,54,6)
    _sc(ws,54,2,'Photos Reference',True,11,ha='center',bdr=_box())
    for row in range(55,59):
        _rh(ws,row,20); _mg(ws,row,2,row,6); _sc(ws,row,2,'',bdr=_box())
    _rh(ws,59,8)

    # ── Technician
    _rh(ws,60,18)
    _sc(ws,60,2,'Technician:',True,10)
    _mg(ws,60,3,60,4); _sc(ws,60,3,r.mechanic or '',sz=10,ul=True)
    _sc(ws,60,5,'Date Completed:',True,10)
    _sc(ws,60,6,str(r.date) if r.date else '',sz=10,ul=True)

    _rh(ws,61,45)
    _sc(ws,61,2,'Signature:',True,10)
    _mg(ws,61,3,61,4)
    if sig_path:
        try:
            sig = XLImg(sig_path)
            sig.width = 120; sig.height = 45; sig.anchor = 'C61'
            ws.add_image(sig)
        except Exception:
            pass

    # ── Footer
    _hline(ws,63,2,6,'thin')
    _rh(ws,64,14); _mg(ws,64,2,64,6)
    _sc(ws,64,2,'www.excelsioraviation.com',sz=9,ha='center',italic=True)
    _outer(ws,1,64,2,6)