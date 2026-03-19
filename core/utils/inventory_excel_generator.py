"""
core/utils/inventory_excel_generator.py
────────────────────────────────────────
Generates an Excel workbook (.xlsx) with two sheets:
  Sheet 1 — Maintenance Service Report Form  (matches SR_00163 PDF layout)
  Sheet 2 — Warranty Certificate             (matches WC2025-015 PDF layout)

Auto-filled from an Inventory model instance.

Usage (in views.py):
    from .utils.inventory_excel_generator import generate_inventory_excel
    buf = generate_inventory_excel(record, logo_path, sig_path)
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImg

DARK_BLUE  = "1F3864"
LIGHT_BLUE = "D9E1F2"
WHITE      = "FFFFFF"
BLACK      = "000000"
GRAY       = "D9D9D9"


# ── Style helpers ─────────────────────────────────────────────────────────────

def _T(style='thin', color='000000'):
    return Side(style=style, color=color)

def _box(thick=False):
    s = _T('medium' if thick else 'thin')
    return Border(top=s, bottom=s, left=s, right=s)

def _bot():
    return Border(bottom=_T('thin'))

def _hline(ws, row, c1, c2, style='medium'):
    s = _T(style)
    for c in range(c1, c2+1):
        cell = ws.cell(row=row, column=c)
        old  = cell.border
        cell.border = Border(top=old.top, bottom=s, left=old.left, right=old.right)

def _sc(ws, row, col, val='', bold=False, sz=10, bg=None, color=BLACK,
        ha='left', va='center', bdr=None, wrap=False, italic=False, ul=False):
    c = ws.cell(row=row, column=col)
    c.value = val
    c.font  = Font(name='Arial', bold=bold, size=sz, color=color,
                   italic=italic, underline='single' if ul else None)
    c.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
    if bg:  c.fill   = PatternFill('solid', start_color=bg)
    if bdr: c.border = bdr

def _mg(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

def _cw(ws, col, w):
    ws.column_dimensions[get_column_letter(col)].width = w

def _rh(ws, row, h):
    ws.row_dimensions[row].height = h

def _outer(ws, r1, r2, c1, c2):
    for r in range(r1, r2+1):
        for col in [c1, c2]:
            cell = ws.cell(row=r, column=col)
            old  = cell.border
            if col == c1:
                cell.border = Border(top=old.top, bottom=old.bottom,
                                     left=_T('medium'), right=old.right)
            else:
                cell.border = Border(top=old.top, bottom=old.bottom,
                                     left=old.left, right=_T('medium'))
    for c in range(c1, c2+1):
        for row in [r1, r2]:
            cell = ws.cell(row=row, column=c)
            old  = cell.border
            if row == r1:
                cell.border = Border(top=_T('medium'), bottom=old.bottom,
                                     left=old.left, right=old.right)
            else:
                cell.border = Border(top=old.top, bottom=_T('medium'),
                                     left=old.left, right=old.right)


def _add_image(ws, path, anchor, w, h):
    """Safely add image — skip if file missing."""
    if not path:
        return
    try:
        img = XLImg(path)
        img.width = w; img.height = h; img.anchor = anchor
        ws.add_image(img)
    except Exception:
        pass


# ── Main generator ────────────────────────────────────────────────────────────

def generate_inventory_excel(record, logo_path=None, sig_path=None):
    """
    Args:
        record    : Inventory model instance
        logo_path : absolute path to logo image (optional)
        sig_path  : absolute path to signature image (optional)
    Returns:
        io.BytesIO buffer ready for HttpResponse
    """
    wb = Workbook()
    _build_msr(wb.active, record, logo_path, sig_path)
    ws2 = wb.create_sheet("Warranty Certificate")
    _build_warranty(ws2, record, logo_path, sig_path)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Sheet 1: Maintenance Service Report Form ──────────────────────────────────

def _build_msr(ws, r, logo_path, sig_path):
    ws.title = "Maintenance Service Report"

    for i, w in enumerate([0.8, 24, 14, 10, 10, 14, 0.8], 1):
        _cw(ws, i, w)
    for row in range(1, 68):
        _rh(ws, row, 15)

    # ── Header
    _rh(ws,2,32); _rh(ws,3,14); _rh(ws,4,14); _rh(ws,5,14)
    _add_image(ws, logo_path, 'B2', 88, 88)

    _mg(ws,2,3,2,6); _sc(ws,2,3,'Excelsior Aviation GSE Corporation',True,15,va='bottom')
    _mg(ws,3,3,3,6); _sc(ws,3,3,'Lot 9 Dona Lucing, Calibutbut Bacolor 2001 Pampanga Philippines',sz=9)
    _mg(ws,4,3,4,6); _sc(ws,4,3,'Mobile No. +63 905 463 7252',sz=9)
    _mg(ws,5,3,5,6); _sc(ws,5,3,'sale.service@excelsioraviation.com     www.excelsioraviation.com',sz=9)

    _rh(ws,6,4); _hline(ws,6,2,6,'medium')
    _rh(ws,7,8)

    # ── Title
    _rh(ws,8,22)
    _mg(ws,8,2,8,6)
    _sc(ws,8,2,'MAINTENANCE SERVICE REPORT FORM',True,13,ha='center',va='center')

    _rh(ws,9,22)
    _mg(ws,9,2,9,6)
    inv_id = f'INV-{r.pk}' if hasattr(r,'pk') and r.pk else 'INV-XXXX'
    _sc(ws,9,2,inv_id,True,12,ha='center',va='center',italic=True,bdr=_box(True))
    _rh(ws,10,10)

    # ── Customer table
    _rh(ws,11,22); _rh(ws,12,22)
    _sc(ws,11,2,'Customer Name',True,10,GRAY,bdr=_box())
    _mg(ws,11,3,11,6); _sc(ws,11,3,r.customer_name or '',sz=10,bdr=_box())
    _sc(ws,12,2,'Address',True,10,GRAY,bdr=_box())
    _mg(ws,12,3,12,6); _sc(ws,12,3,r.customer_address or '',sz=10,bdr=_box(),wrap=True)
    _rh(ws,13,10)

    # ── Equipment table
    for row, lbl, val in [
        (14, 'Equipment Description:', r.description  or ''),
        (15, 'Part No.',               r.model_number or ''),
        (16, 'Serial No.',             r.serial_number or ''),
        (17, 'Capacity / Quantity',    str(r.quantity) if r.quantity else ''),
    ]:
        _rh(ws,row,20)
        _sc(ws,row,2,lbl,True,10,GRAY,bdr=_box())
        _mg(ws,row,3,row,6); _sc(ws,row,3,val,sz=10,bdr=_box())
    _rh(ws,18,10)

    # ── Description of Service
    _rh(ws,19,22)
    _mg(ws,19,2,19,6)
    _sc(ws,19,2,'Description of Service',True,11,ha='center',bdr=_box())
    for i, item in enumerate(['Inspection','Repair','Load Test']):
        _rh(ws,20+i,18)
        _mg(ws,20+i,2,20+i,6)
        _sc(ws,20+i,2,f'•  {item}',sz=10,bdr=_box())
    _rh(ws,23,10)

    # ── Corrective Action
    _rh(ws,24,22)
    _mg(ws,24,2,24,6)
    _sc(ws,24,2,'Describe corrective action completed or required',True,11,ha='center',bdr=_box())

    prefill = [
        f'•  Inspect unit — {r.description or ""}',
        f'•  Serial No. {r.serial_number or ""}  /  Part No. {r.model_number or ""}',
        f'•  Location: {r.location or ""}',
        f'•  Quantity on hand: {r.quantity or ""}',
    ]
    for i in range(10):
        _rh(ws,25+i,18)
        _mg(ws,25+i,2,25+i,6)
        line = prefill[i] if i < len(prefill) else ''
        _sc(ws,25+i,2,line,sz=10,bdr=_box(),wrap=True)
    _rh(ws,35,10)

    # ── Remarks
    _rh(ws,36,22)
    _mg(ws,36,2,36,6)
    _sc(ws,36,2,'Remarks',True,11,ha='center',bdr=_box())
    _rh(ws,37,22)
    _mg(ws,37,2,37,6)
    _sc(ws,37,2,f'•  {r.remarks}' if r.remarks else '',sz=10,bdr=_box(),wrap=True)
    for row in [38,39]:
        _rh(ws,row,18); _mg(ws,row,2,row,6); _sc(ws,row,2,'',bdr=_box())
    _rh(ws,40,10)

    # ── Parts Installed
    _rh(ws,41,22)
    _mg(ws,41,2,41,6)
    _sc(ws,41,2,'Parts Installed',True,11,ha='center',bdr=_box())
    _rh(ws,42,20)
    for col, hdr in zip([2,3,4,5,6],['Qty','Unit','Description','','Part Number']):
        _sc(ws,42,col,hdr,True,10,GRAY,ha='center',bdr=_box())
    for i in range(6):
        _rh(ws,43+i,18)
        for col in range(2,7): _sc(ws,43+i,col,'',bdr=_box())
    _rh(ws,49,10)

    # ── Photos Reference
    _rh(ws,50,22)
    _mg(ws,50,2,50,6)
    _sc(ws,50,2,'Photos reference',True,11,ha='center',bdr=_box())
    for row in range(51,58):
        _rh(ws,row,22); _mg(ws,row,2,row,6); _sc(ws,row,2,'',bdr=_box())
    _rh(ws,58,10)

    # ── Technician
    _rh(ws,59,20)
    tech = r.created_by.get_full_name() if hasattr(r,'created_by') and r.created_by else ''
    date = r.created_at.strftime('%d %B %Y') if hasattr(r,'created_at') and r.created_at else ''
    _sc(ws,59,2,'Technician:',True,10)
    _mg(ws,59,3,59,4); _sc(ws,59,3,tech,sz=10,ul=True)
    _sc(ws,59,5,'Date Completed:',True,10)
    _sc(ws,59,6,date,sz=10,ul=True)

    _rh(ws,60,50)
    _sc(ws,60,2,'Signature:',True,10)
    _mg(ws,60,3,60,4)
    _add_image(ws, sig_path, 'C60', 120, 48)

    _hline(ws,62,2,6,'thin')
    _rh(ws,63,16)
    _mg(ws,63,2,63,6)
    _sc(ws,63,2,'www.excelsioraviation.com',sz=9,ha='center',italic=True)
    _outer(ws,1,63,2,6)


# ── Sheet 2: Warranty Certificate ────────────────────────────────────────────

def _build_warranty(ws, r, logo_path, sig_path):
    for i, w in enumerate([0.8,11,10,10,10,10,10,10,10,0.8], 1):
        _cw(ws, i, w)
    for row in range(1, 38): _rh(ws, row, 15)

    # ── Header
    _rh(ws,2,32); _rh(ws,3,14); _rh(ws,4,14); _rh(ws,5,14)
    _add_image(ws, logo_path, 'B2', 88, 88)

    _mg(ws,2,3,2,9); _sc(ws,2,3,'Excelsior Aviation GSE Corporation',True,15,va='bottom')
    _mg(ws,3,3,3,9); _sc(ws,3,3,'Lot 9 Dona Lucing, Calibutbut Bacolor 2001 Pampanga Philippines',sz=9)
    _mg(ws,4,3,4,9); _sc(ws,4,3,'Mobile No. +63 905 463 7252  +63 917 158 8844',sz=9)
    _mg(ws,5,3,5,9); _sc(ws,5,3,'sale.service@excelsioraviation.com     www.excelsioraviation.com',sz=9)
    _rh(ws,6,4); _hline(ws,6,2,9,'medium')
    _rh(ws,7,8)

    # Cert No
    _rh(ws,8,16)
    _mg(ws,8,2,8,9)
    year = r.created_at.year if hasattr(r,'created_at') and r.created_at else '2026'
    _sc(ws,8,2,f'No. WC{year}-XXXX',True,11,ha='right',ul=True)

    # Title
    _rh(ws,9,28)
    _mg(ws,9,2,9,9)
    _sc(ws,9,2,'WARRANTY CERTIFICATE',True,16,ha='center',va='center')
    _rh(ws,10,4); _hline(ws,10,2,9,'thin')
    _rh(ws,11,8)

    # ── Fields
    for row in range(12,18): _rh(ws,row,18)

    _sc(ws,12,2,'CUSTOMER',True,10);  _sc(ws,12,3,':',sz=10)
    _mg(ws,12,4,12,6); _sc(ws,12,4,r.customer_name or '',bdr=_bot())
    _sc(ws,12,7,'ADDRESS',True,10);   _sc(ws,12,8,':',sz=10)
    _sc(ws,12,9,r.customer_address or '',bdr=_bot(),wrap=True)

    date_str = r.created_at.strftime('%B %d, %Y') if hasattr(r,'created_at') and r.created_at else ''
    _sc(ws,13,2,'DATE OF SERVICE',True,10); _sc(ws,13,3,':',sz=10)
    _mg(ws,13,4,13,6); _sc(ws,13,4,date_str,bdr=_bot())
    _sc(ws,13,7,'WARRANTY PERIOD',True,10); _sc(ws,13,8,':',sz=10)
    _sc(ws,13,9,'2 Years',bdr=_bot())

    _sc(ws,14,2,'ITEM / DESCRIPTION',True,10); _sc(ws,14,3,':',sz=10)
    _mg(ws,14,4,14,9); _sc(ws,14,4,r.description or '',bdr=_bot())

    _sc(ws,15,2,'MODEL NUMBER',True,10); _sc(ws,15,3,':',sz=10)
    _mg(ws,15,4,15,6); _sc(ws,15,4,r.model_number or '',bdr=_bot())
    _sc(ws,15,7,'QUANTITY',True,10); _sc(ws,15,8,':',sz=10)
    _sc(ws,15,9,str(r.quantity) if r.quantity else '',bdr=_bot())

    _sc(ws,16,2,'SERIAL NUMBER',True,10); _sc(ws,16,3,':',sz=10)
    _mg(ws,16,4,16,6); _sc(ws,16,4,r.serial_number or '',bdr=_bot())
    _sc(ws,16,7,'LOCATION',True,10); _sc(ws,16,8,':',sz=10)
    _sc(ws,16,9,r.location or '',bdr=_bot())

    _rh(ws,17,8)

    # ── Warranty Statement
    _rh(ws,18,16)
    _mg(ws,18,2,18,9)
    _sc(ws,18,2,'WARRANTY STATEMENT',True,11)

    _rh(ws,19,55)
    _mg(ws,19,2,19,9)
    desc = r.description or 'The described equipment'
    _sc(ws,19,2,
        f'{desc} is fully warranted against Excelsior technician faults and defective parts '
        f'under our service warranty guide for a period of 2 Years from the date of service '
        f'and delivery or upon completion of installation.',
        sz=10, wrap=True, va='top')
    _rh(ws,20,8)

    # ── Exceptions
    _rh(ws,21,16)
    _mg(ws,21,2,21,9)
    _sc(ws,21,2,'WARRANTY EXCEPTIONS:',True,10,LIGHT_BLUE,BLACK)

    exceptions = [
        "1.  Damage resulting from accidents, misuse, abuse, over loading and failure to follow normal operating procedures outlined in the user's manual.",
        "2.  Normal wear-and-tear, corrosion, rusting or stains.",
        "3.  Defects and damage arising from improper testing, operation, demonstration, maintenance, installation, adjustment or any alteration or modification of any kind.",
        "4.  General maintenance and routine servicing made other than Excelsior Aviation GSE Corporation authorized technician.",
        "5.  If any parts of the unit are replaced (third party brand) not supplied or approved by Excelsior or unit been dismantled and repair by the other technician.",
    ]
    for i, exc in enumerate(exceptions):
        row = 22+i
        _rh(ws,row, 28 if len(exc)>100 else 18)
        _mg(ws,row,2,row,9)
        _sc(ws,row,2,exc,sz=10,bdr=_box(),wrap=True,va='top')
    _rh(ws,27,8)

    # ── Signature
    _rh(ws,28,52); _rh(ws,29,16); _rh(ws,30,16)
    _add_image(ws, sig_path, 'B28', 120, 48)

    _mg(ws,28,6,28,9); _sc(ws,28,6,date_str,sz=10,ha='right',va='bottom')
    tech = r.created_by.get_full_name() if hasattr(r,'created_by') and r.created_by else ''
    _mg(ws,29,2,29,5); _sc(ws,29,2,tech,sz=10,ul=True)
    _mg(ws,29,6,29,9); _sc(ws,29,6,'Date',sz=10,ha='right')
    _mg(ws,30,2,30,5); _sc(ws,30,2,'Technician',sz=10,italic=True)
    _rh(ws,31,8)

    # Notes
    _rh(ws,32,45)
    _mg(ws,32,2,32,9)
    _sc(ws,32,2,
        'Excelsior Aviation GSE Corporation shall not be liable of any incidents due to human error, '
        'item is been tested and meet the rated load limit on the date of service. '
        'Certificate is not valid without Excelsior dry seal.',
        sz=9, italic=True, wrap=True, va='top')

    # Footer
    _hline(ws,34,2,9,'thin')
    _rh(ws,35,16)
    _mg(ws,35,2,35,9)
    _sc(ws,35,2,'www.excelsioraviation.com',sz=9,ha='center',italic=True)
    _outer(ws,1,35,2,9)