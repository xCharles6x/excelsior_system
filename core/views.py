from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Q
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator
import io

from .models import LoadTest, Inventory, RepairInspection, Certificate, Customer, TestEquipmentRow, LoadCellCertRow, LoadCellSavedCert
from .forms import (LoadTestForm, InventoryForm, RepairInspectionForm,
                    CertificateForm, TestEquipmentFormSet, LoadCellCertFormSet,
                    CustomerForm, CreateUserForm)



# ── HELPERS ───────────────────────────────────────────────────────────────────

def admin_required(view_func):
    """Decorator: login required + must be staff/superuser."""
    @login_required
    def wrapper(request, *args, **kwargs):
        if not request.user.is_staff:
            messages.error(request, 'Admin access required.')
            return redirect('dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper


# ── AUTH ──────────────────────────────────────────────────────────────────────

def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        user = authenticate(request,
                            username=request.POST.get('username'),
                            password=request.POST.get('password'))
        if user:
            login(request, user)
            return redirect('dashboard')
        messages.error(request, 'Invalid username or password.')
    return render(request, 'login.html')


def logout_view(request):
    logout(request)
    return redirect('login')


# ── DASHBOARD ─────────────────────────────────────────────────────────────────

@login_required
def dashboard(request):
    return render(request, 'dashboard.html')


# ── LOAD TEST ─────────────────────────────────────────────────────────────────

@login_required
def loadtest(request):
    if request.method == 'POST':
        form = LoadTestForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.created_by = request.user
            obj.save()
            messages.success(request, 'Load Test record saved successfully!')
            return redirect('loadtest')
        messages.error(request, 'Please correct the errors below.')
    else:
        form = LoadTestForm()
    return render(request, 'loadtest.html', {'form': form})


# ── INVENTORY ─────────────────────────────────────────────────────────────────

@login_required
def inventory(request):
    if request.method == 'POST':
        form = InventoryForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.created_by = request.user
            obj.save()
            messages.success(request, 'Inventory record saved successfully!')
            return redirect('inventory')
        messages.error(request, 'Please correct the errors below.')
    else:
        form = InventoryForm()
    return render(request, 'inventory.html', {'form': form})


# ── REPAIR ────────────────────────────────────────────────────────────────────

@login_required
def repair(request):
    if request.method == 'POST':
        form = RepairInspectionForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.created_by = request.user
            obj.save()
            messages.success(request, 'Repair record saved successfully!')
            return redirect('repair')
        messages.error(request, 'Please correct the errors below.')
    else:
        form = RepairInspectionForm(initial={'record_type': 'repair'})
    return render(request, 'repair.html', {'form': form, 'page_title': 'Repair'})


# ── INSPECTION ────────────────────────────────────────────────────────────────

@login_required
def inspection(request):
    if request.method == 'POST':
        form = RepairInspectionForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.record_type = 'inspection'
            obj.created_by = request.user
            obj.save()
            messages.success(request, 'Inspection record saved successfully!')
            return redirect('inspection')
        messages.error(request, 'Please correct the errors below.')
    else:
        form = RepairInspectionForm(initial={'record_type': 'inspection'})
    return render(request, 'repair.html', {'form': form, 'page_title': 'Inspection'})


# ── MASTER LIST ───────────────────────────────────────────────────────────────

@login_required
def list_records(request):
    record_type  = request.GET.get('type', 'loadtest')
    search_query = request.GET.get('search', '').strip()

    loadtest_records  = LoadTest.objects.all()
    inventory_records = Inventory.objects.all()
    repair_records    = RepairInspection.objects.all()

    if search_query:
        if record_type == 'loadtest':
            loadtest_records = loadtest_records.filter(
                Q(customer_name__icontains=search_query) |
                Q(company__icontains=search_query) |
                Q(description__icontains=search_query) |
                Q(serial_number__icontains=search_query) |
                Q(certificate_number__icontains=search_query) |
                Q(mechanic__icontains=search_query)
            )
        elif record_type == 'inventory':
            inventory_records = inventory_records.filter(
                Q(customer_name__icontains=search_query) |
                Q(description__icontains=search_query) |
                Q(model_number__icontains=search_query) |
                Q(serial_number__icontains=search_query) |
                Q(location__icontains=search_query)
            )
        elif record_type == 'repair':
            repair_records = repair_records.filter(
                Q(customer_name__icontains=search_query) |
                Q(company__icontains=search_query) |
                Q(description__icontains=search_query) |
                Q(serial_number__icontains=search_query) |
                Q(report_number__icontains=search_query) |
                Q(mechanic__icontains=search_query)
            )

    def paginate(qs):
        p = Paginator(qs, 20)
        return p.get_page(request.GET.get('page', 1))

    context = {
        'record_type':       record_type,
        'search_query':      search_query,
        'loadtest_records':  paginate(loadtest_records)  if record_type == 'loadtest'  else loadtest_records,
        'inventory_records': paginate(inventory_records) if record_type == 'inventory' else inventory_records,
        'repair_records':    paginate(repair_records)    if record_type == 'repair'    else repair_records,
        'loadtest_count':    loadtest_records.count(),
        'inventory_count':   inventory_records.count(),
        'repair_count':      repair_records.count(),
    }
    return render(request, 'list.html', context)


# ── EDIT VIEWS ────────────────────────────────────────────────────────────────

@login_required
def edit_loadtest(request, pk):
    record = get_object_or_404(LoadTest, pk=pk)
    if request.method == 'POST':
        form = LoadTestForm(request.POST, instance=record)
        if form.is_valid():
            form.save()
            messages.success(request, 'Load Test record updated successfully!')
            return redirect('/list/?type=loadtest')
        messages.error(request, 'Please correct the errors below.')
    else:
        form = LoadTestForm(instance=record)
    return render(request, 'edit_record.html', {
        'form': form, 'record_type': 'loadtest',
        'record': record, 'page_title': 'Edit Load Test',
    })


@login_required
def edit_inventory(request, pk):
    record = get_object_or_404(Inventory, pk=pk)
    if request.method == 'POST':
        form = InventoryForm(request.POST, instance=record)
        if form.is_valid():
            form.save()
            messages.success(request, 'Inventory record updated successfully!')
            return redirect('/list/?type=inventory')
        messages.error(request, 'Please correct the errors below.')
    else:
        form = InventoryForm(instance=record)
    return render(request, 'edit_record.html', {
        'form': form, 'record_type': 'inventory',
        'record': record, 'page_title': 'Edit Inventory',
    })


@login_required
def edit_repair(request, pk):
    record = get_object_or_404(RepairInspection, pk=pk)
    if request.method == 'POST':
        form = RepairInspectionForm(request.POST, instance=record)
        if form.is_valid():
            form.save()
            messages.success(request, 'Record updated successfully!')
            return redirect(f'/list/?type={record.record_type}')
        messages.error(request, 'Please correct the errors below.')
    else:
        form = RepairInspectionForm(instance=record)
    return render(request, 'edit_record.html', {
        'form': form, 'record_type': record.record_type,
        'record': record, 'page_title': f'Edit {record.get_record_type_display()}',
    })


# ── DELETE VIEWS ──────────────────────────────────────────────────────────────

@login_required
@require_POST
def delete_loadtest(request, pk):
    get_object_or_404(LoadTest, pk=pk).delete()
    messages.success(request, 'Load Test record deleted.')
    return redirect('/list/?type=loadtest')


@login_required
@require_POST
def delete_inventory(request, pk):
    get_object_or_404(Inventory, pk=pk).delete()
    messages.success(request, 'Inventory record deleted.')
    return redirect('/list/?type=inventory')


@login_required
@require_POST
def delete_repair(request, pk):
    record = get_object_or_404(RepairInspection, pk=pk)
    rtype  = record.record_type
    record.delete()
    messages.success(request, 'Record deleted.')
    return redirect(f'/list/?type={rtype}')


# ── EXPORT ────────────────────────────────────────────────────────────────────

@login_required
def export_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return HttpResponse('openpyxl not installed.', status=500)

    record_type  = request.GET.get('type', 'loadtest')
    search_query = request.GET.get('search', '').strip()

    wb          = openpyxl.Workbook()
    ws          = wb.active
    header_font  = Font(bold=True, color='1C1C2E')
    header_fill  = PatternFill('solid', fgColor='F5A623')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin         = Side(style='thin', color='CCCCCC')
    thin_border  = Border(left=thin, right=thin, top=thin, bottom=thin)

    if record_type == 'loadtest':
        ws.title = 'Load Test Records'
        headers = ['Customer', 'Customer Address', 'Company', 'Description',
                   'Model No.', 'Serial No.', 'Equipment',
                   'Tested Load', 'Safe Load', 'Certificate No.',
                   'Remark', 'Mechanic', 'Date Inspection', 'Date Due']
        qs = LoadTest.objects.all()
        if search_query:
            qs = qs.filter(
                Q(customer_name__icontains=search_query) |
                Q(company__icontains=search_query) |
                Q(description__icontains=search_query)
            )
        ws.append(headers)
        for r in qs:
            ws.append([r.customer_name, r.customer_address, r.company, r.description,
                       r.model_number, r.serial_number, r.equipment,
                       r.tested_load, r.safe_load, r.certificate_number,
                       r.remark, r.mechanic,
                       str(r.date_inspection or ''), str(r.date_due or '')])

    elif record_type == 'inventory':
        ws.title = 'Inventory Records'
        headers = ['Customer', 'Customer Address', 'Description',
                   'Model No.', 'Serial No.', 'Location', 'Quantity', 'Remarks']
        qs = Inventory.objects.all()
        if search_query:
            qs = qs.filter(
                Q(customer_name__icontains=search_query) |
                Q(description__icontains=search_query) |
                Q(serial_number__icontains=search_query)
            )
        ws.append(headers)
        for r in qs:
            ws.append([r.customer_name, r.customer_address, r.description,
                       r.model_number, r.serial_number, r.location,
                       r.quantity, r.remarks])

    elif record_type == 'repair':
        ws.title = 'Repair & Inspection Records'
        headers = ['Customer', 'Customer Address', 'Type', 'Company', 'Description',
                   'Model No.', 'Serial No.', 'Report No.',
                   'Customer Report', 'Diagnose Result', 'Remarks', 'Date', 'Mechanic']
        qs = RepairInspection.objects.all()
        if search_query:
            qs = qs.filter(
                Q(customer_name__icontains=search_query) |
                Q(company__icontains=search_query) |
                Q(description__icontains=search_query)
            )
        ws.append(headers)
        for r in qs:
            ws.append([r.customer_name, r.customer_address,
                       r.get_record_type_display(), r.company, r.description,
                       r.model_number, r.serial_number, r.report_number,
                       r.customer_report, r.diagnose_result, r.remarks,
                       str(r.date or ''), r.mechanic])

    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = thin_border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border    = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"excelsior_{record_type}_records.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ── CERTIFICATE VIEWS ─────────────────────────────────────────────────────────

@login_required
def customer_address_api(request):
    cid = request.GET.get('id')
    try:
        c = Customer.objects.get(pk=cid)
        return JsonResponse({'address': c.address})
    except Customer.DoesNotExist:
        return JsonResponse({'address': ''})


@login_required
def certificate_list(request):
    certs = Certificate.objects.select_related('customer').all()
    q = request.GET.get('search', '').strip()
    if q:
        certs = certs.filter(
            Q(certificate_number__icontains=q) |
            Q(customer__name__icontains=q) |
            Q(product_description__icontains=q)
        )
    paginator = Paginator(certs, 20)
    page = paginator.get_page(request.GET.get('page', 1))
    return render(request, 'certificate_list.html', {
        'certificates': page,
        'search_query': q,
    })


@login_required
def certificate_create(request):
    if request.method == 'POST':
        form    = CertificateForm(request.POST)
        formset = TestEquipmentFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            cert = form.save()
            formset.instance = cert
            formset.save()
            messages.success(request, f'Certificate {cert.certificate_number} created successfully.')
            return redirect('certificate_list')
        messages.error(request, 'Please correct the errors below.')
    else:
        form    = CertificateForm()
        formset = TestEquipmentFormSet()
    return render(request, 'certificate_form.html', {
        'form': form, 'formset': formset,
        'lc_all_certs': LoadCellSavedCert.objects.all(),
        'page_title': 'New Certificate', 'is_edit': False,
    })


@login_required
def certificate_edit(request, pk):
    cert = get_object_or_404(Certificate, pk=pk)
    if request.method == 'POST':
        form    = CertificateForm(request.POST, instance=cert)
        formset = TestEquipmentFormSet(request.POST, instance=cert)
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            messages.success(request, f'Certificate {cert.certificate_number} updated.')
            return redirect('certificate_list')
        messages.error(request, 'Please correct the errors below.')
    else:
        form    = CertificateForm(instance=cert)
        formset = TestEquipmentFormSet(instance=cert)
    return render(request, 'certificate_form.html', {
        'form': form, 'formset': formset,
        'lc_all_certs': LoadCellSavedCert.objects.all(),
        'cert': cert,
        'page_title': f'Edit Certificate {cert.certificate_number}',
        'is_edit': True,
    })


@login_required
def certificate_download(request, pk):
    from .utils.certificate_generator import generate_certificate
    cert = get_object_or_404(Certificate, pk=pk)
    rows = cert.test_equipment_rows.all()
    buf  = generate_certificate(cert, rows)
    filename = f'Certificate_{cert.certificate_number}.docx'
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@require_POST
def certificate_delete(request, pk):
    cert = get_object_or_404(Certificate, pk=pk)
    num  = cert.certificate_number
    cert.delete()
    messages.success(request, f'Certificate {num} deleted.')
    return redirect('certificate_list')


@login_required
def customer_create(request):
    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            c = form.save()
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'id': c.pk, 'name': c.name, 'address': c.address})
            messages.success(request, f'Customer "{c.name}" added.')
            return redirect('certificate_create')
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'error': 'Invalid data'}, status=400)
    else:
        form = CustomerForm()
    return render(request, 'customer_form.html', {'form': form})


# ── LOAD CELL SAVED LIST API ─────────────────────────────────────────────────

@csrf_exempt
@login_required
def loadcell_cert_save(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed.'}, status=405)
    cert_number = request.POST.get('cert_number', '').strip()
    if not cert_number:
        return JsonResponse({'error': 'Certificate number is required.'}, status=400)
    obj, created = LoadCellSavedCert.objects.get_or_create(cert_number=cert_number)
    if not created:
        return JsonResponse({'error': 'Already exists in the list.'}, status=400)
    return JsonResponse({'cert_number': obj.cert_number})


@csrf_exempt
@login_required
def loadcell_cert_delete(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed.'}, status=405)
    cert_number = request.POST.get('cert_number', '').strip()
    if not cert_number:
        return JsonResponse({'error': 'Certificate number is required.'}, status=400)
    LoadCellSavedCert.objects.filter(cert_number=cert_number).delete()
    return JsonResponse({'success': True})


# ── ADMIN: USER MANAGEMENT ────────────────────────────────────────────────────

@admin_required
def admin_user_list(request):
    """Admin page: list all user accounts."""
    users = User.objects.all().order_by('username')
    return render(request, 'admin_users.html', {'users': users})


@admin_required
def admin_user_create(request):
    """Admin page: create a new user account."""
    if request.method == 'POST':
        form = CreateUserForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.set_password(form.cleaned_data['password'])
            user.save()
            messages.success(request, f'Account "{user.username}" created successfully!')
            return redirect('admin_user_list')
        messages.error(request, 'Please correct the errors below.')
    else:
        form = CreateUserForm()
    return render(request, 'admin_user_form.html', {
        'form': form,
        'page_title': 'Create Account',
        'action_label': 'Create Account',
    })


@admin_required
def admin_user_edit(request, pk):
    """Admin page: edit an existing user account."""
    target = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        form = CreateUserForm(request.POST, instance=target)
        if form.is_valid():
            user = form.save(commit=False)
            # Only update password if a new one was provided
            new_pw = form.cleaned_data.get('password')
            if new_pw:
                user.set_password(new_pw)
                if request.user == user:
                    update_session_auth_hash(request, user)
            user.save()
            messages.success(request, f'Account "{user.username}" updated.')
            return redirect('admin_user_list')
        messages.error(request, 'Please correct the errors below.')
    else:
        form = CreateUserForm(instance=target)
    return render(request, 'admin_user_form.html', {
        'form': form,
        'target_user': target,
        'page_title': f'Edit Account: {target.username}',
        'action_label': 'Save Changes',
    })


@admin_required
@require_POST
def admin_user_delete(request, pk):
    """Admin: delete a user account."""
    target = get_object_or_404(User, pk=pk)
    if target == request.user:
        messages.error(request, 'You cannot delete your own account.')
        return redirect('admin_user_list')
    username = target.username
    target.delete()
    messages.success(request, f'Account "{username}" deleted.')
    return redirect('admin_user_list')

@login_required
@require_POST
def customer_delete(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        if customer.certificates.exists():
            return JsonResponse({
                'success': False,
                'error': f'Cannot delete "{customer.name}" because they have existing certificates. Delete the certificates first.'
            })
        name = customer.name
        customer.delete()
        return JsonResponse({'success': True, 'name': name})
    # Non-AJAX fallback
    if customer.certificates.exists():
        messages.error(request, f'Cannot delete "{customer.name}" - they have existing certificates.')
        return redirect('certificate_create')
    name = customer.name
    customer.delete()
    messages.success(request, f'Customer "{name}" deleted.')
    return redirect('certificate_create')


# ── GO TO CERTIFICATION (pre-fill from records) ───────────────────────────

@login_required
def cert_from_loadtest(request, pk):
    """Pre-fill certificate form using LoadTest record data."""
    record = get_object_or_404(LoadTest, pk=pk)

    # Find or create Customer
    customer = None
    if record.customer_name:
        customer, _ = Customer.objects.get_or_create(
            name=record.customer_name,
            defaults={'address': record.customer_address or ''}
        )

    initial = {
        'customer':           customer.pk if customer else None,
        'address':            record.customer_address or '',
        'product_description': record.description or '',
        'capacity':           record.safe_load or '',
        'model_number':       record.model_number or '',
        'serial_number':      record.serial_number or '',
        'tested_load':        record.tested_load or '',
        'working_load_limits': record.safe_load or '',
        'date_of_inspection': record.date_inspection or '',
        'due_next_inspection': record.date_due or '',
        'technician_name':    record.mechanic or '',
    }

    form       = CertificateForm(initial=initial)
    formset    = TestEquipmentFormSet()
    lc_formset = LoadCellCertFormSet()
    return render(request, 'certificate_form.html', {
        'form':       form,
        'formset':    formset,
        'lc_formset': lc_formset,
        'page_title': 'New Certificate',
        'is_edit':    False,
        'lc_all_certs': LoadCellSavedCert.objects.all(),
        'prefilled_from': f'Load Test record — {record.description or record.pk}',
    })


@login_required
def cert_from_repair(request, pk):
    """Pre-fill certificate form using Repair/Inspection record data."""
    record = get_object_or_404(RepairInspection, pk=pk)

    customer = None
    if record.customer_name:
        customer, _ = Customer.objects.get_or_create(
            name=record.customer_name,
            defaults={'address': record.customer_address or ''}
        )

    initial = {
        'customer':            customer.pk if customer else None,
        'address':             record.customer_address or '',
        'product_description': record.description or '',
        'model_number':        record.model_number or '',
        'serial_number':       record.serial_number or '',
        'date_of_inspection':  record.date or '',
        'technician_name':     record.mechanic or '',
    }

    form       = CertificateForm(initial=initial)
    formset    = TestEquipmentFormSet()
    lc_formset = LoadCellCertFormSet()
    return render(request, 'certificate_form.html', {
        'form':       form,
        'formset':    formset,
        'lc_formset': lc_formset,
        'page_title': 'New Certificate',
        'is_edit':    False,
        'lc_all_certs': LoadCellSavedCert.objects.all(),
        'prefilled_from': f'{record.get_record_type_display()} record — {record.description or record.pk}',
    })


@login_required
def cert_from_inventory(request, pk):
    """Pre-fill certificate form using Inventory record data."""
    record = get_object_or_404(Inventory, pk=pk)

    customer = None
    if record.customer_name:
        customer, _ = Customer.objects.get_or_create(
            name=record.customer_name,
            defaults={'address': record.customer_address or ''}
        )

    initial = {
        'customer':            customer.pk if customer else None,
        'address':             record.customer_address or '',
        'product_description': record.description or '',
        'model_number':        record.model_number or '',
        'serial_number':       record.serial_number or '',
        'capacity':            str(record.quantity) if record.quantity else '',
    }

    form       = CertificateForm(initial=initial)
    formset    = TestEquipmentFormSet()
    lc_formset = LoadCellCertFormSet()
    return render(request, 'certificate_form.html', {
        'form':       form,
        'formset':    formset,
        'lc_formset': lc_formset,
        'page_title': 'New Certificate',
        'is_edit':    False,
        'lc_all_certs': LoadCellSavedCert.objects.all(),
        'prefilled_from': f'Inventory record — {record.description or record.pk}',
    })